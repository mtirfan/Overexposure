import driver
import view
import sys
import graph_creation
import create_graph_from_file as cff
import networkx as nx
import openpyxl
import re
FL_PREFIX = "test_files/"
FILE_DIRECTORY_PREFIX = "currently_in_use/"

def is_ID_reset():
    filename = FILE_DIRECTORY_PREFIX + "ID_tracker.txt"
    file = open(filename, 'r')
    # get current ID that is recorded in file
    if file.mode == 'r':
        ID = file.read()

    return ID == "1"

def generate_and_save_graphs(node_sizes, runs, m=2,p=0.2):
    if not is_ID_reset():
        do_stop = input("The ID counter is not reset, is this intentional? Data may be overwritten if you continue. Enter [Y] or [N]: ")
        if do_stop == "Y":
            sys.exit()
    for num_nodes in node_sizes:
        for i in range(runs):
            original_graphs, original_types = driver.generate_original_graphs(num_nodes, m, p)
            for O, graph_type in zip(original_graphs, original_types):
                ID = view.generate_ID()
                location = view.save_original(O, "n/a", "n/a", graph_type, ID, "n/a", "n/a")

def test_dif_combos(node_sizes, runs, k_vals, appeals):
    graph_types = ["WS"] # ["BA","ER",
    offset = [3] # [1,2,

    for graph, ID_off in zip(graph_types, offset):
        for appeal in appeals:
            for num_nodes_offset, num_nodes in enumerate(node_sizes):
                for k in k_vals:
                    for i in range(runs):
                        filename = graph + "/" + str(num_nodes) + "/" + str((num_nodes_offset * runs * 3) + (i * 3) + ID_off)
                        driver.test_file(filename, k, appeal)

def get_all_possible_cluster_sizes(node_sizes, runs, appeals):
    graph_types = ["BA","ER","WS"]
    offset = [1,2,3]
    
    possible_cluster_sizes = set()
    for graph, ID_off in zip(graph_types, offset):
        for appeal in appeals:
            appeal = float(appeal)
            for num_nodes_offset, num_nodes in enumerate(node_sizes):
                num_nodes_offset = 3
                for i in range(runs):
                    filename = "test_files/" + graph + "/" + str(num_nodes) + "/" + str((num_nodes_offset * runs * 3) + (i * 3) + ID_off) + ".txt"
                    print("Opening " + filename)
                    file_k, file_appeal, graph_type, ID, file_remove_cycles, file_assumption_1, O = cff.create_from_file(filename)
                    C = graph_creation.create_cluster_graph(O, appeal)
                    B = graph_creation.create_bipartite_from_cluster(C)
                    for node in B.nodes():
                        if B.nodes[node]['bipartite'] == 0:
                            possible_cluster_sizes.add(B.nodes[node]['weight'])
                    print("Finished recording clusters for " + filename)
                    B.clear()
                    C.clear()
                    O.clear()
    possible_cluster_sizes = list(possible_cluster_sizes)
    possible_cluster_sizes.sort(reverse=True)
    print("======================================")
    print("There are " + str(len(possible_cluster_sizes)) + " unique cluster sizes")

    # record all cluster sizes into a file
    filename = FILE_DIRECTORY_PREFIX + "all_cluster_sizes.txt"
    # put number on new line everytime
    with open(filename, 'w') as write_to_file:
        for size in possible_cluster_sizes:
            write_to_file.write(str(size) + "\n")
    print("Finished recording in file")

def record_occurences_of_cluster_sizes(node_sizes, runs, appeals):
    file = open(FILE_DIRECTORY_PREFIX + "all_cluster_sizes.txt","r")
    if file.mode == 'r':
        contents = file.read()
    lines = re.split("\n", contents)
    file.close()

    graph_types = ["BA","ER","WS"]
    offset = [1,2,3]

    write_header(lines)
    
    for graph, ID_off in zip(graph_types, offset):
        for appeal in appeals:
            appeal = float(appeal)
            for num_nodes_offset, num_nodes in enumerate(node_sizes):
                num_nodes_offset = 3
                for i in range(runs):
                    cluster_sizes = reset_dict_from_list_keys(lines)
                    filename = "test_files/" + graph + "/" + str(num_nodes) + "/" + str((num_nodes_offset * runs * 3) + (i * 3) + ID_off) + ".txt"
                    print("Opening " + filename)
                    file_k, file_appeal, graph_type, ID, file_remove_cycles, file_assumption_1, O = cff.create_from_file(filename)
                    C = graph_creation.create_cluster_graph(O, appeal)
                    B = graph_creation.create_bipartite_from_cluster(C)
                    for node in B.nodes():
                        if B.nodes[node]['bipartite'] == 0:
                            if B.nodes[node]['weight'] not in cluster_sizes.keys():
                                print("ENCOUNTERED A WEIGHT NOT IN DICTIONARY")
                                sys.exit()
                            cluster_sizes[B.nodes[node]['weight']] = cluster_sizes[B.nodes[node]['weight']] + 1
                    print("Finished Count for " + filename)
                    save_cluster_size_data("5000", graph, appeal, filename, cluster_sizes)
                    B.clear()
                    C.clear()
                    O.clear()

def reset_dict_from_list_keys(lines):
    cluster_sizes = dict()
    for line in lines:
        if line != '':
            cluster_sizes[int(line)] = 0
    return cluster_sizes  

def save_cluster_size_data(num_nodes, graph_type, appeal, filename, cluster_sizes):
    print("Svaing Data")
    wb = openpyxl.load_workbook(FILE_DIRECTORY_PREFIX + 'Cluster_Data.xlsx')
    sheets = wb.sheetnames
    cluster_size_data = wb[sheets[0]]
    counts = [value for key, value in cluster_sizes.items()]
    cluster_size_data.append([num_nodes] + [appeal] + [graph_type] + [filename] + counts)
    wb.save(FILE_DIRECTORY_PREFIX + 'Cluster_Data.xlsx')

def write_header(lines):
    print("Write Header")
    wb = openpyxl.load_workbook(FILE_DIRECTORY_PREFIX + 'Cluster_Data.xlsx')
    sheets = wb.sheetnames
    cluster_size_data = wb[sheets[0]]
    cluster_size_data.append(["Num Nodes","Appeal","Graph Type","Location"] + lines)
    wb.save(FILE_DIRECTORY_PREFIX + 'Cluster_Data.xlsx')

def histogram_data_collection(node_sizes, runs, appeals):
    get_all_possible_cluster_sizes(node_sizes, runs, appeals)
    record_occurences_of_cluster_sizes(node_sizes, runs, appeals)

def main():
    node_sizes = [500, 1000, 2000, 5000] # ["500", "1000", "2000", "5000"] | ["150", "500", "1000", "2000"]
    k_vals = ["10","20","50","100"] # ["5","10","20","50"]
    appeals = ["0.75"] # ["0.5", "0.5", "0.5", "0.5"]
    runs = 25

    m = 2
    p = 0.2

    # if not is_ID_reset():
    #     do_gen = input("Do you want to generate new files? Enter [Y] or [N]: ")
    #     if do_gen == "Y":
    #         generate_and_save_graphs(node_sizes, runs, m=2,p=0.2)

    test_dif_combos(node_sizes, runs, k_vals, appeals)

    # histogram_data_collection([5000], runs, appeals)

main()











# get_cluster_data()
# driver.test_file("1", "5", "0.5")

# # the functions below were lazy ways to generate and test programs, relying on generation while testing
# def generate_all_possible_combos():
#     node_sizes = ["500", "1000", "2000", "5000"] # ["150", "500", "1000", "2000"]
#     possible_k = ["10","20","50","100"] # ["5","10","20","50"]
#     possible_criticalities = ["0.5", "0.75"] # ["0.5", "0.5", "0.5", "0.5"]

#     do_remove_cycles = "False"
#     do_assumption_1 = "False"

#     runs = 25

#     for criticality in possible_criticalities:
#         for num_nodes in node_sizes:
#             for k in possible_k:
#                 for i in range(runs):
#                     print("RUN " + str(i) + ": " + str(num_nodes) + " " + str(k) + " " + str(criticality))
#                     driver.test_new_file(num_nodes, k, criticality, do_remove_cycles, do_assumption_1)

# def test():
#     graph_types = ["BA"]
#     offset = [1]
#     node_size = ["150"]
#     do_remove_cycles = "False"
#     do_assumption_1 = "False"

#     for graph, ID_off in zip(graph_types, offset):
#         for size_offset, size in enumerate(node_size):
#             size_offset = 0
#             for i in range(50):
#                 filename = graph + "/" + size + "/" + str((size_offset * 150) + (i * 3) + ID_off)
#                 # if filename[:3] != "BA/":
#                 driver.retest_old_file(filename, do_remove_cycles, do_assumption_1)

# def practice():
#     num_nodes = "500"
#     k = "10"
#     criticality = "0.5"
#     do_remove_cycles = "False"
#     do_assumption_1 = "False"

#     for i in range(6):
#         print("RUN " + str(i) + ": " + str(num_nodes) + " " + str(k) + " " + str(criticality))
#         driver.test_new_file(num_nodes, k, criticality, do_remove_cycles, do_assumption_1)
# def retest_all_ER_WS():
#     graph_types = ["ER","WS"]
#     offset = [2,3]
#     node_size = ["150", "500", "1000", "2000"]
#     do_remove_cycles = "False"
#     do_assumption_1 = "False"

#     for graph, ID_off in zip(graph_types, offset):
#         for size_offset, size in enumerate(node_size):
#             for i in range(50):
#                 filename = graph + "/" + size + "/" + str((size_offset * 150) + (i * 3) + ID_off)
#                 # if filename[:3] != "BA/":
#                 driver.retest_old_file(filename, do_remove_cycles, do_assumption_1)

# def retest_all_files():
#     graph_types = ["BA","ER","WS"]
#     offset = [1,2,3]
#     node_size = ["150", "500", "1000", "2000"]
#     do_remove_cycles = "False"
#     do_assumption_1 = "False"

#     for graph, ID_off in zip(graph_types, offset):
#         for size_offset, size in enumerate(node_size):
#             for i in range(50):
#                 filename = graph + "/" + size + "/" + str((size_offset * 150) + (i * 3) + ID_off)
#                 # if filename[:3] != "BA/":
#                 driver.retest_old_file(filename, do_remove_cycles, do_assumption_1)

# def retest_BA():
#     graph_types = ["BA"]
#     offset = [1]
#     node_size = ["150", "500", "1000", "2000"]
#     do_remove_cycles = "False"
#     do_assumption_1 = "False"

#     for graph, ID_off in zip(graph_types, offset):
#         for size_offset, size in enumerate(node_size):
#             for i in range(50):
#                 filename = graph + "/" + size + "/" + str((size_offset * 150) + (i * 3) + ID_off)
#                 # if filename[:3] != "BA/":
#                 driver.retest_old_file(filename, do_remove_cycles, do_assumption_1)

# # practice()