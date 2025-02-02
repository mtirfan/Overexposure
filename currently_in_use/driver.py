"""
Driver for entire program.
How to run the program:
The program takes four commad line arguments. 
The first is the program name, the second is the number of 
nodes in the original graph, the third is the number of seeds
we are choosing for our seed set, and the fourth is the criticality
parameter. For example, typing the following in the terminal would 
output an optimal seed set of size 5 for an original graph with 50 nodes
and a criticality threshold of 0.7 (any node whose criticality is above
0.7 will be an accepting node)

python3 driver.py 50 5 0.7

Note that the above line requires that you are inside the currently_in_use folder

The main method which reads in these arguments passes them onto runTests(),
which acts as a driver for the following algorithms:

1) greedy dynamic programming
2) greedy knapsack
3) recursive DP (optimal for trees where <=2 clusters share any rej node)
4) Brute force (commented out by default because it takes so long. User
can choose to uncomment it and run)
5) Linear Program (optimal only when <= 2 clusters share any rej node)
6) Bipartite linear program (optimal on any graph)

All of the results are printed to the terminal and then
written to an excel sheet which is within the folder currently_in_use/tests

We also print the original, cluster, and bipartite graph to be used for comparison.
Note that the first thing we do in runTests is try to create 
a graph that satisfies our requirements, namely:
    1) No rejecting node is shared by more than 2 clusters
    2) No cycles in the cluster graph

We continuously try to make the cluster graph until these
properties are satisfied.
TODO: find a more efficient way to satisfy these properties. For 
original graphs with >100 nodes, it takes a long time to create
a cluster graph which satisfies the properties.
"""
import create_clusters as cc
import DP_algorithms as dp
import greedy_approx_algorithms as greedy
import brute_force as bf
import bipartite_linear_program as blp
import create_graph_from_file as cff
import networkx as nx
import matplotlib.pyplot as plt
import make_bipartite_graph as mbg
import sys
import openpyxl
import xlwt
import cluster_linear_program as clp
import timeit
from datetime import datetime
import networkx as nx
import bipartite_approx_algs as baa
import synthetic_graph_creation as sgc

# use "currently_in_use/" if in Overexposue folder, "" if in currently_in_use already (personal war im fighting with the vs code debugger)
FILE_DIRECTORY_PREFIX = ""#"currently_in_use/"

c = 0.5
n = 10 #num nodes
m = 2 #num links
p = 0.2 #prob rewrite edge

#TODO: run more rigorous tests
#TODO: fix logging output to take max of root and no root
#TODO: work on paper, write analysis on run times/approximation factor
#TODO: tree decomposition algorithm

"""
Experiment Design:
    - previous experiments in previous papers
    - what kinds of networks they are building, how many nodes they are taking into account
    - SNAP graphs?
    - we are limited to cluster graphs
    - Erdos-Renyi
    - Barabasi-Alber
    - Watts-Strogatz
    - real world networks like college messaging

Come up with an experiment design based on papers that we've read, then see how they are doing their experiments section.
    What are the common things they are doing? What kinds of random graph structures?
    Outline what the things they are showing are
    Write some brief notes and then design the experiment and then go with that
    Present those things next week
    We want to save the graph and run tests based on that
"""

def get_graph(num_nodes, k, criticality):
    G_cluster = False
    G = False
    cluster_graphs, original_graphs = sgc.make_graphs(criticality, num_nodes)
    graph_types = ["ba-no-cycle","ba-cycle", "er-no-cycle","er-cycle", "ws-no-cycle","ws-cycle"]
    while G_cluster == False:
        G = nx.random_tree(num_nodes)
        G_cluster = cc.testOriginaltoCluster(G, num_nodes, criticality, True, True)
    cluster_graphs.append(G_cluster)
    store_info(G_cluster, k)
    graph_types.append("cluster no cycle")
    G_cluster_cycle = False
    while G_cluster_cycle == False:
        G_cluster_cycle = cc.testOriginaltoCluster(G, num_nodes, criticality, False, False)
    cluster_graphs.append(G_cluster_cycle)
    original_graphs.append(["tree", G])
    graph_types.append("cluster cycle")
    for original in original_graphs:
        cc.showOriginalGraph(original[1], criticality)
        plt.savefig(FILE_DIRECTORY_PREFIX + "saved-graphs/"+original[0] + ".png")
        # plt.show()
    return cluster_graphs, graph_types
    
    """
    if graph_type == "1":
        print("making graph")
        while G_cluster == False:
            G = nx.random_tree(num_nodes)
            G_cluster = cc.testOriginaltoCluster(G, num_nodes, criticality, True)
            return [G_cluster], [G]
    elif graph_type == "2":
        while G_cluster == False:
            G = nx.random_tree(num_nodes)
            G_cluster = cc.testOriginaltoCluster(G, num_nodes, criticality, False)
            return [G_cluster], [G]
    elif graph_type == "3" or graph_type == "4" or graph_type == "5":
        return cluster_graphs, original_graphs
    
    elif graph_type == "4":
        while G_cluster == False:
            G = nx.barabasi_albert_graph(num_nodes, 3)
            G_cluster = cc.testOriginaltoCluster(G, num_nodes, criticality, True)
    elif graph_type == "5":
        while G_cluster == False:
            G = nx.watts_strogatz_graph(num_nodes, 3, 0.5)
            G_cluster = cc.testOriginaltoCluster(G, num_nodes, criticality, True)
    else:
        print("Try again, no graph type specified or spelled wrong")
        exit(0)
    """

"""
Driver for all of our algorithms.
@params:
    num_nodes --> number of nodes in the original graph
    k --> number of seeds chosen. Corresponds to number of clusters
    picked in the cluster graph
    criticality --> criticality threshold for nodes in the original graph.
    Nodes above this value will be accepting
"""
def runTests(num_nodes, k, criticality):
    #create cluster graph
    G_clusters, graph_types = get_graph(num_nodes, k, criticality) #initialize to false, when we get a graph that satisfies the requirements
    for graph in G_clusters:
        print("NODES: ", graph.nodes())
    #this will be true
    #TODO: commented out. Used when we want to create a cluster graph without an original graph
    #max_weight = 5
    #TODO: commented out. Comment out the above 3 lines and uncomment this to create a tree cluster
    #graph without starting with the original. Useful for testing.
    #G = cc.createClusterGraph(num_nodes, max_weight)
    for G, graph_type in zip(G_clusters, graph_types):
        #compute payoff for greedy DP
        print("NEXT TEST:------------->", graph_type)
        start = timeit.default_timer()
        max_val_greedyDP, seedset = greedy.greedyDP(G, G.number_of_nodes(), k)
        stop = timeit.default_timer()
        runtime_greedy_DP = stop - start

        store_info(G,k)
        print("\nGreedy DP Payoff: ", max_val_greedyDP)
            
        #compute payoff for most basic greedy algorithm
        start = timeit.default_timer()
        greedy_payoff, greedy_seedset = greedy.kHighestClusters(G, k)
        stop = timeit.default_timer()
        runtime_greedy = stop - start
        print("Greedy Approach Seeds Chosen:", greedy_seedset, " with payoff: ", greedy_payoff)

        #compute payoff for recursive DP
        start = timeit.default_timer()
        payoff_root, payoff_no_root = dp.runRecursiveDP(G, k)
        payoff_recursive_dp = max(payoff_root, payoff_no_root)
        stop = timeit.default_timer()
        runtime_recursive_DP = stop - start
        print("Recursive DP payoff: ", payoff_recursive_dp)

        #run linear program
        start = timeit.default_timer()
        payoff_clp = clp.lp_setup(G, k)
        stop = timeit.default_timer()
        runtime_cluster_LP = stop - start

        #run bipartite linear program
        bipartite = mbg.graph_to_bipartite(G)
        start = timeit.default_timer()
        payoff_blp = blp.solve_lp(bipartite, k)
        stop = timeit.default_timer()
        runtime_bipartite_LP = stop - start

        start = timeit.default_timer()
        payoff_greedy = baa.greedy_selection(bipartite, k)
        stop = timeit.default_timer()

        #compute payoff using brute force algorithm --> uncomment out if you want to run
        #best_payoff_selection,best_payoff = bf.computePayoff(bipartite, k)
        #print("Brute Force payoff: ", best_payoff_selection, best_payoff)

        runtime_greedy_bipartite = stop - start

        start = timeit.default_timer()
        payoff_forward_thinking = baa.forward_thinking_greedy(bipartite, k)
        stop = timeit.default_timer()
        runtime_forward_thinking = stop - start

        #write the results to excel file
        if graph_type == "ba-cycle" or graph_type == "ws-cycle" or graph_type == "er-cycle" or graph_type == "cluster cycle":
            write_results("-","-", "-", "-", "-", \
                "-", "-", payoff_blp, payoff_greedy, payoff_forward_thinking, "-","-", "-", "-",runtime_bipartite_LP, \
                runtime_greedy_bipartite, runtime_forward_thinking, num_nodes,k, graph_type, criticality)
        else:
            write_results(max_val_greedyDP,greedy_payoff, payoff_recursive_dp, payoff_clp, payoff_blp, \
                payoff_greedy, payoff_forward_thinking, "-", "-", "-", runtime_greedy_DP, runtime_greedy, runtime_recursive_DP, \
                runtime_cluster_LP, runtime_bipartite_LP, runtime_greedy_bipartite, runtime_forward_thinking, num_nodes,k, graph_type, criticality)
        #print cluster graph and bipartite graph
        #cc.showOriginalGraph(original, criticality)
        printGraph(G, graph_type)
        printBipartite(bipartite, graph_type)
    # plt.show()

""" Print bipartite graph using network x. Saved to file"""
def printBipartite(bipartite, name):
    print("printing bipartite graph")
    plt.figure(name+ "-bipartite")
    pos = nx.spring_layout(bipartite)
    nx.draw(bipartite, pos)

    node_labels = nx.get_node_attributes(bipartite,'weight')
    print(node_labels)
    # do (id, weight) pair for lable instead of just weight
    for key,val in node_labels.items():
        node_labels[key] = (key,val)
    nx.draw_networkx_labels(bipartite, pos=pos, labels=node_labels)
    plt.savefig("saved-graphs/"+ name + "-bipartite.png")

""" display cluster graph """
def printGraph(G, name):
    plt.figure(name)
    pos = nx.spring_layout(G)
    nx.draw(G, pos)

    node_labels = nx.get_node_attributes(G,'weight')
    # do (id, weight) pair for lable instead of just weight
    for key,val in node_labels.items():
        node_labels[key] = (key,val)
    nx.draw_networkx_labels(G, pos=pos, labels=node_labels)
    data_info = nx.get_edge_attributes(G,'rej_nodes') # edge lables rejecting node
    weight_info = nx.get_edge_attributes(G,'weight') # edge lables rejecting node
    edge_labels = {}
    for key in weight_info.keys():
        if key in data_info.keys():
            edge_labels[key] = (data_info[key],weight_info[key])
        else:
            edge_labels[key] = ("na",weight_info[key])
    nx.draw_networkx_edge_labels(G, pos=pos, edge_labels=edge_labels)
    plt.savefig("saved-graphs/" + name + ".png")

""" Store more specific info about each graph, to be used for testing if the 
results output in the excel sheet are inaccurate / do not make sense """
def store_info(G,k):
    print('\nNext Test:\n')
    with open(FILE_DIRECTORY_PREFIX + "tests/cluster_graph_details.txt", 'w') as graph_info:
        timestamp = datetime.timestamp(datetime.now())
        date = datetime.fromtimestamp(timestamp)
        graph_info.write('c\n')
        graph_info.write("# Prints each node weight on a new line followed by each edge. \n# \
Ie the first number printed is the weight of node 0. Then prints each edge followed by its weight and the rejecting nodes.\n# \
For example, 0 2 1 -22 indicates that there is an edge (0,2) with weight 1, and the rejecting node in that edge is node 22. \n")
        graph_info.write("# Timestamp: " + str(date) + "\n")
        graph_info.write("# Nodes: " + str(G.number_of_nodes()) + "\n")
        data = G.edges.data()
        graph_info.write("# Edges: " + str(len(data)))
        weights = G.nodes.data('weight')
        for node in weights:
            graph_info.write("\n" + str(node[1]))
        for item in data:
            graph_info.write("\n" + str(item[0]) + " " + str(item[1]) + " " + str(item[2]['weight']))
            try:
                data = item[2]['rej_nodes']
                for reject in data:
                    graph_info.write(" " + str(reject))
            except:
                pass

""" Write results to an excel sheet stored in the currently_in_use/tests folder """
def write_results(max_val_greedyDP,greedy_payoff, payoff_recurisve_dp, payoff_clp, payoff_blp, \
    payoff_bipartite_greedy, payoff_forward_thinking, payoff_blp_cycles, payoff_bipartite_greedy_cycles, payoff_forward_thinking_cycles, \
    runtime_greedy_DP, runtime_greedy, runtime_recursive_DP, \
    runtime_cluster_LP, runtime_bipartite_LP, runtime_greedy_bipartite, runtime_forward_thinking, n, k, graph_type, criticality):
    wb = openpyxl.load_workbook('tests/Test_results.xlsx')
    print("WRITING RESULTS")
    sheets = wb.sheetnames
    data = wb[sheets[0]]
    runtimes = wb[sheets[1]]
    timestamp = datetime.timestamp(datetime.now())
    date = str(datetime.fromtimestamp(timestamp))
    row = data.max_row+1
    row2 = runtimes.max_row+1
    data_items_to_add = [n, k, criticality, graph_type, date, max_val_greedyDP, greedy_payoff, payoff_recurisve_dp, payoff_clp, payoff_blp, payoff_bipartite_greedy, payoff_forward_thinking, payoff_blp_cycles, payoff_bipartite_greedy_cycles, payoff_forward_thinking_cycles]
    runtime_data_to_add = [n, k, criticality, graph_type, date, runtime_greedy_DP, runtime_greedy, runtime_recursive_DP, \
    runtime_cluster_LP, runtime_bipartite_LP, runtime_greedy_bipartite, runtime_forward_thinking]
    i = 1
    j = 1
    for item in data_items_to_add:
        c1 = data.cell(row = row, column = i)
        c1.value = item
        i += 1
    for item in runtime_data_to_add:
        c1 = runtimes.cell(row = row2, column = j)
        c1.value = item
        j += 1
    
    wb.save('tests/Test_results.xlsx')

def getUserInput():
    """
    graph_type = input("What type of graph would you like to run tests on? Options are:\n 1. Cluster graph \
with no cycles such that no rejecting node shared by more than 2 clusters. Type 1 \n 2. Cluster graph \
such that no rejecting node shared by more than 2 clusters (cycles may be present). Type 2.\n 3.\
Erdos-Renyi . Type 3 \n 4. Barabasi-Alber. Type 4 \n 5. Watts-Strogatz. Type 5\n")
    """
    graph_type = "all"
    return graph_type

#main function, used for calling things
def main(num_seeds, k, criticality):
    num_seeds = int(num_seeds)
    k = int(k)
    criticality = float(criticality)
    print("getting user input")
    #graph_type = getUserInput()
    runTests(num_seeds, k, criticality)

if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], sys.argv[3])

# crits = [0.4,0.5,0.7]
# nn = 75
# times = 2
# kk = 5
# for crit in crits:
#     for i in range(times):
#         main(nn, kk, crit)